import openpyxl

# creating a new excel workbook via openpyxl.Workbook()
wk = openpyxl.Workbook()  # workbook object wk
sh = wk.active  # assigning the active worksheet to sh object
sh.title = "Testing"  # changing the active sheet title to Testing. Default is sheet
print(sh.title)
sh['A2'].value = "Testing World"  # writing data into a particular cell of the active excel sheet

# creating another sheet in the existing WriteExcel.xlsx file
wk.create_sheet(title="Testing2")
sh1 = wk['Testing2']
sh1['A2'] = '91-9804401899'  # writing data in the A2 cell of Testing2 sheet

# remove any sheet
# wk.remove_sheet(wk['Testing']) # it will throw error : DeprecationWarning: Call to deprecated function remove_sheet
# (Use wb.remove(worksheet) or del wb[sheetname]). but still it will work
# instead use below to remove any sheet , and no error will come
# wk.remove(wk['Testing']) # while removing any sheet either pass the sheet object name or else pass the sheet title
wk.remove(sh)

wk.save("..\\TestData\\WriteExcel.xlsx")  # saving the excel workbook
