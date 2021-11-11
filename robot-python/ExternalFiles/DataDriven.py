import openpyxl

wk = openpyxl.load_workbook("../TestData/datadriventesting.xlsx")


# finding out max data of rows]
def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row

#finding out the cell data
def fetch_cell_data(Sheetname, row, col):
    sh = wk[Sheetname]
    cell_data = sh.cell(int(row),int(col))
    return cell_data.value


# print(fetch_number_of_rows("Sheet1"))
# print(fetch_cell_data("Sheet1", 1, 1))
