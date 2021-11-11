# import package to work with excel
import openpyxl

# Load Workbook
wk = openpyxl.load_workbook("..\\TestData\\Sample.xlsx")
print(wk.sheetnames)  # get the diff sheetnames.o/p:- ['Sheet1', 'Sheet2', 'Sheet3']
print("Active sheet is: " + wk.active.title)  # fetch the title of the active sheet .o/p:-Active sheet is: Sheet1

# # # Read one cell data # # #
# create object of any sheet in which you want to work
sh1 = wk['Sheet1']  # creating object of Sheet 1
print(sh1.title)  # o/p :- Sheet1 #prints the title of the active sheet
# fetching cell data via sheet object
print(sh1['A3'].value)  # get the A3 cell value of sheet1.o/p:- Username3
print(sh1['B4'].value)  # o/p:- Password4
# another way of fetching cell data using cell object
cell1 = sh1.cell(1, 1)
print(cell1.value)  # o/p :- Username1
cell2 = sh1.cell(column=1, row=3)
print(cell2.value)  # o/p:- Username3
# to find out for which column and row is the cell1 object created
print(cell1.row)  # o/p:- 3
print(cell1.column)  # o/p:- 1

# # # Read all rows and cells data ###
# first fetch the max row no and col no
rows = sh1.max_row  # max_row fetches the max row no of the sheet object
columns = sh1.max_column  # max_column fetches the max column no of the sheet object

print('Total rows are: ' + str(rows))  # o/p :- Total rows are: 5
print('Total columns are: ' + str(columns))  # o/p:- Total columns are: 3

# traverse through each cell using the for loop and the sheet object
for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        c = sh1.cell(i, j)
        print(c.value)

# Another way of fetching all cells data by mentioning the cell value in the for loop and traversing through each cell

for r in sh1['A1':'C5']:
    for c in r:
        print(c.value)
